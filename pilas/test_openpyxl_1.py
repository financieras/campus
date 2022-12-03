import openpyxl                # importamos la librería openpyxl
wb = openpyxl.Workbook()       # para crear por primera vez un libro
ws = wb.active                 # estas dos lineas crean el libro, aún sin nombre

print(type(wb))                # miramos a ver quién es la variable wb
print(wb.sheetnames)           # proporciona la lista de hojas del libro de trabajo
for i in range(2,11):          # creamos las hojas Hoja2 hasta Hoja10
    wb.create_sheet("Sheet"+str(i))

ws1 = wb['Sheet']              # cambiamos el nombre de la primera hoja llamada Sheet
ws1.title = 'Sheet1'           # por el nombre Sheet1

print(wb.sheetnames)           # imprimimos todos los nombres de hoja
wb.save('mejor_opcion.xlsx')   # grabamos el fichero por primera vez