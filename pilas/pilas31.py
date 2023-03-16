# FUNCIONES
def sa(a, b, result):   # result es un array que va acumulando los pasos dados. Ejemplo: result = ['pb','ra','sa','rra']
    if len(a) > 1: a[0],a[1] = a[1],a[0]; result.append('sa')
    return a, b, result
def sb(a, b, result):
    if len(b) > 1: b[0],b[1] = b[1],b[0]; result.append('sb')
    return a, b, result
def ss(a, b, result):   # ss no llama a sa y sb ya que en result quedarían anotadas las tres funciones: ss, sa y sb
    if len(a) > 1 or len(b) > 1:
        result.append('ss')
        if len(a) > 1: a[0],a[1] = a[1],a[0]
        if len(b) > 1: b[0],b[1] = b[1],b[0]
    return a, b, result
def pa(a, b, result):
    if len(b) > 0:
        a.insert(0, b[0])
        b.pop(0); result.append('pa')
    return a, b, result
def pb(a, b, result):
    if len(a) > 0:
        b.insert(0, a[0])
        a.pop(0); result.append('pb')
    return a, b, result
def ra(a, b, result):
    if len(a) > 1: a.append(a.pop(0)); result.append('ra')
    return a, b, result
def rb(a, b, result):
    if len(b) > 1: b.append(b.pop(0)); result.append('rb')
    return a, b, result
def rr(a, b, result):
    if len(a) > 1 or len(b) > 1:
        result.append('rr')
    if len(a) > 1: a.append(a.pop(0))
    if len(b) > 1: b.append(b.pop(0))
    return a, b, result
def rra(a, b, result):
    if len(a) > 1: a.insert(0, a.pop()); result.append('rra')
    return a, b, result
def rrb(a, b, result):
    if len(b) > 1: b.insert(0, b.pop()); result.append('rrb')
    return a, b, result
def rrr(a, b, result):
    if len(a) > 1 or len(b) > 1:
        result.append('rrr')
    if len(a) > 1: a.insert(0, a.pop())
    if len(b) > 1: b.insert(0, b.pop())
    return a, b, result


def establece_lis(arr):   # esta función solo se debería invocar una vez. Proporciona una matriz con listas de posibles lis
    m2 = [[0,1],[0,-1],[0,2],[0,-2],[0,3],[0,-3],[1,-1],[1,2],[1,-2],[1,3],[1,-3],[-1,2],[-1,-2],[-1,3],[-1,-3],[2,-2],[2,3],[2,-3],[-2,3],[-2,-3],[3,-3]]
    posibles_lis = []
    for pareja in m2:   # m2 tiene 21 parejas
        x,y = pareja[0], pareja[1]
        lis = [arr[x], arr[y]]
        posibles_lis.append(lis)
    return posibles_lis

# Llevamos de la pila A hacia la pila B los elementos de la LIS
def rotateLIS(a, b, result, lis):
    for vlis in lis:
        if a.index(vlis) < len(a)/2:
            for i in range(a.index(vlis)):
                ra(a, b, result)
        else:
            for i in range(len(a)- a.index(vlis)):
                rra(a, b, result)
        pb(a, b, result)



# PASOS NECESARIOS PARA COLOCAR CADA ELEMENTO DE A EN SU SITIO EN B      total = pasosA + pasosB (esta suma es solo una idea, se ha de sumar de una forma peculiar)

def necesariosA(a, b):   # array pasosA calcula los pasos necesarios para colocar cada elemento de A como el primero de su pila
    for v in a:
        if a.index(v) < len(a)/2:
            pasosA.append(a.index(v))
        else:
            pasosA.append(-(len(a)- a.index(v)))

def necesariosB(a, b):   # array pasosB calcula los pasos de B necesarios para colocar cada elemento de A dentro de su sitio en B
    for i in range(len(a)):    # objetivo_primero es el número que se ha de poner en la primera posición de la pila B
        if a[i] < min(b):      # si el elemento de A considerado es menor que el mayor de B entonces
            objetivo_primero = max(b)   # el objetivo_primero será el mayor de B
        else:   # objetivo_primero en este caso será el maximo de los inferiores en B al valor iésimo de A
            objetivo_primero = min(b)   #se inicializa en el valor mínimo de la pila B
            for j in range(len(b)):
                if b[j] < a[i] and b[j] > objetivo_primero:
                        objetivo_primero = b[j]
        # el objetivo_primero se ha de situar el primero de la pila B            
        if b.index(objetivo_primero) < len(b)/2:
            pasosB.append(b.index(objetivo_primero))
        else:
            pasosB.append(-(len(b)- b.index(objetivo_primero)))

def totaliza(a, b):   # totalizar pasos
    global total
    for i in range(len(pasosA)):
        if pasosA[i] * pasosB[i] < 0:   # si son de distinto signo, uno positivo y otro negativo
            total.append(abs(pasosA[i]) + abs(pasosB[i]))   # no hay sinergia
        else:   # si son de igual signo o alguno cero
            total.append(max(abs(pasosA[i]), abs(pasosB[i])))  # si son de igual signo hay sinergia

def calculaIndexPasosMinimos():
    global pasosA, pasosB, total
    pasosA = []   # [0,1,2, ...,41,-41,... , -2,-1]  vector donde cada index está asociado con el valor del mismo index en A
    pasosB = []   # [-3,2,-5, ..., -5,0]   estos dos arrays se han de recalcular cada vez que realmente se mueva algún elemento de A a B
    total = []
    necesariosA(a, b)
    necesariosB(a, b)
    totaliza(a, b)
    return total.index(min(total))   # retorna el índice del elemento de la pila A que menos pasos totales necesita



def giraA(a, b, indice, steps, result):
    for i in range(abs(steps)):
        if steps > 0:
            ra(a, b, result)
        elif steps < 0:
            rra(a, b, result)

def giraB(a, b, indice, steps, result):
    for i in range(abs(steps)):
        if steps > 0:
            rb(a, b, result)
        elif steps < 0:
            rrb(a, b, result)

# girando pilas A y B
def giraPilas(a, b, indice, result):   # indice es el index del valor en la pila A que deseamos poner el primero
    if pasosA[indice] * pasosB[indice] > 0:   # Existe sinergia, nos podemos ahorrar pasos
        pasos_comunes = min(abs(pasosA[indice]), abs(pasosB[indice]))
        for i in range(pasos_comunes):
            if pasosA[indice] > 0:   # si el signo de ambos es positivo, ya que ambos tienen el mismo signo
                rr(a, b, result)
            elif pasosA[indice] < 0: # si el signo de ambos es negativo
                rrr(a, b, result)
        exceso_pasosA = abs(pasosA[indice]) - pasos_comunes
        exceso_pasosB = abs(pasosB[indice]) - pasos_comunes
        giraA(a, b, indice, ((pasosA[indice] > 0) - (pasosA[indice] < 0)) * exceso_pasosA, result)    # (a > 0) - (a < 0) da el signo de a
        giraB(a, b, indice, ((pasosB[indice] > 0) - (pasosB[indice] < 0)) * exceso_pasosB, result)   # Python no tiene función sign
    else:   # No existe sinergia
        giraA(a ,b, indice, pasosA[indice], result)   # gira A
        giraB(a, b, indice, pasosB[indice], result)   # gira B

    


def situarMax_en_B(a, b, result):   # situa el valor máximo de B en la primera posición de B
    indice = b.index(max(b))
    if indice < len(a)/2:
        steps = indice
    else:
        steps = -(len(b)- indice)
    giraB(a, b, indice, steps, result)

def crecientesA(a, b, result):   # situa los tres valores de A de la forma adecuada según el caso
    if a[0] < a[1] > a[2] and a[0] < a[2]:          # caso: 1 3 2
        ra(a, b, result)
        sa(a, b, result)
        rra(a, b, result)
    elif a[0] > a[1] < a[2] and a[0] < a[2]:        # caso: 2 1 3
        sa(a, b, result)
    elif a[0] < a[1] > a[2] and a[0] > a[2]:        # caso: 2 3 1
        rra(a, b, result)
    elif a[0] > a[1] < a[2] and a[0] > a[2]:        # caso: 3 1 2
        ra(a, b, result)
    elif a[0] > a[1] > a[2]:                        # caso: 3 2 1
        sa(a, b, result)
        rra(a, b, result)
    return                                          # caso: 1 2 3 en este caso no se hace nada pq ya están ordenados



# Después de ordenar los tres elementos de A en orden creciente estricto con la función crecientesA

def cremallera(a, b, result):   # función que va pasando los números de B a A en forma de cremallera, intercalándolos con los que ya hay en A
    aux = [a[-3], a[-2], a[-1]] # creamos un array aux con los tres últimos valores de A
    while len(aux) > 0:         # Mientras aux tenga elementos
        maximo = max(b + aux)   # calcula el maximo entre b y aux
        if maximo in b:         # si el maximo está en B hacer pa
            pa(a, b, result)
        if maximo in aux:       # si el máximo está en aux hacer rra y aux.pop
            rra(a, b, result)
            aux.pop()
    while len(b) > 0:           # ahora aux está vacío y solo queda hacer pa todo el rato
        pa(a, b, result)



if __name__ == "__main__":
    from random import seed, shuffle
    import openpyxl                # importamos la librería openpyxl

    wb = openpyxl.Workbook()       # para crear por primera vez un libro
    ws = wb.active                 # estas dos lineas crean el libro, aún sin nombre

    wb.save('mapa_parejas.xlsx')          # grabamos el fichero por primera vez
    seed()
    n = 500                             # número de elementos de la pila
    for muestra in range(681):
        a = list(range(1, n+1)); shuffle(a) # generación aleatoria de la pila A
        b = []
        a_original = a[:]
        posibles_lis = establece_lis(a)     # llama a la función que proporciona una matriz con las posibles lis de dos números de la pila A
        mejor_contador = 32767              # calcularemos los mínimos pasos necesarios de los intentos. Se inicializa en infinito o casi
        mejor_result = []                   # contendrá el array con el result de la mejor solución que correspone a la de minimos pasos requeridos
        ws = wb["Sheet"]                    # accediendo a la hoja Sheet
        ws.cell(row=muestra+11, column=2).value = muestra + 1
        ws.cell(row=muestra+11, column=3).value = str(a_original)
        for intento in range(21):           # vamos a probar cada una de las 21 parejas. Cada prueba será un intento
            a = a_original[:]
            b = []
            result = []                     # recoje los pasos dados en este intento. Ejemplo: result = ['pb','ra','sa','rra']
            lis = posibles_lis[intento]     # establece lis por ejemplo: lis = [a[0], a[1]]
            rotateLIS(a, b, result, lis)    # pasa de A a B la pareja de elementos de A elegidos en la lis
            while len(a) > 3:
                #print(f"\n{a}\n{b}")
                index_minimosPasos = calculaIndexPasosMinimos()
                giraPilas(a, b, index_minimosPasos, result)
                pb(a, b, result)         # lo pasa de A a B haciendo pb
            situarMax_en_B(a, b, result) # situa el valor máximo de B en la primera posición de B
            crecientesA(a, b, result)    # hace que los tres últimos elementos de A queden ordenados en forma creciente estricta
            cremallera(a, b, result)     # función que va pasando los números de B a A en forma de cremallera, intercalándolos con los que ya hay en A
            nresult = len(result)
            ws.cell(row=muestra+11, column=intento+4).value = nresult
        wb.save('mapa_parejas.xlsx')
    wb.close()