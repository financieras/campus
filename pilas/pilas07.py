# Creación de un archivo en Excel con la solución aportada por nuestro algoritmo
# Se aplica a las permutaciones de un variado número de elementos
from random import sample, seed
import openpyxl            # importamos la librería openpyxl
from itertools import permutations
seed()

import openpyxl                # importamos la librería openpyxl
wb = openpyxl.Workbook()       # para crear por primera vez un libro
ws = wb.active                 # estas dos lineas crean el libro, aún sin nombre

print(type(wb))                # miramos a ver quién es la variable wb
print(wb.sheetnames)           # proporciona la lista de hojas del libro de trabajo
for i in range(2,11):          # creamos las hojas Hoja2 hasta Hoja10
    wb.create_sheet("Sheet"+str(i))

ws1 = wb['Sheet']              # cambiamos el nombre de la primera hoja llamada Sheet
ws1.title = 'Sheet1'           # por el nombre Sheet1

print(wb.sheetnames)           # imprimimos todos los nombres de hoja
wb.save('pilas_algoritmo.xlsx')   # grabamos el fichero por primera vez


###  Trabajando con un archivo Excel concreto  ###
# Cuando ya sebemos que queremos trabajar con el archivo campus/pilas/pilas_algoritmo.xlsm
#wb = openpyxl.load_workbook('pilas_algoritmo.xlsx')

ws4 = wb["Sheet4"]         # accediendo a la hoja4
ws4['E10'] = "Permutaciones"

n = 4
perm = permutations(list(range(1,n+1)))   # obtenemos todas las permutaciones de la lista 

for count, fila in enumerate(list(perm)):             # imprimimos todas las permutaciones
    lista = list(fila)                                # las primera fila es una tupla con (1,2,3,4)
    for j in range(n):
        ws4.cell(row=count+11, column=j+5).value = lista[j]

wb.save('mejor_opcion.xlsx')
wb.close()
