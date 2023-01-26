# código que no funciona correctamente, tiene errores. Ver el pilas7.py

from random import sample, seed
import openpyxl            # importamos la librería openpyxl
from itertools import permutations
seed()

def lista_ordenada(l):
    estaordenada = False
    aux = l[:]
    aux.sort()
    if (aux == l):
        estaordenada = True
    return estaordenada

def sa(a,b):
    if len(a) > 1: a[0],a[1] = a[1],a[0]
    return a,b
def sb(a,b):
    if len(b) > 1: b[0],b[1] = b[1],b[0]
    return a,b
def ss(a,b):   # luego no se usa
    sa(a,b)
    sb(a,b)
    return a,b
def pa(a,b):
    if len(b) > 0:
        a.insert(0, b[0])
        b.pop(0)
    return a,b
def pb(a,b):
    if len(a) > 0:
        b.insert(0, a[0])
        a.pop(0)
    return a,b
def ra(a,b):
    if len(a) > 1: a.append(a.pop(0))
    return a,b
def rb(a,b):
    if len(b) > 1: b.append(b.pop(0))
    return a,b
def rr(a,b):   # luego no se usa
    ra(a,b)
    rb(a,b)
    return a,b
def rra(a,b):
    if len(a) > 1: a.insert(0, a.pop())
    return a,b
def rrb(a,b):
    if len(b) > 1: b.insert(0, b.pop())
    return a,b
def rrr(a,b):   # luego no se usa
    rra(a,b)
    rrb(a,b)
    return a,b

if __name__ == "__main__":
    # Cuando ya sebemos que queremos trabajar con el archivo campus/pilas/mejor_opcion.xlsm
    wb = openpyxl.load_workbook('mejor_opcion.xlsx')
    n = 4
    ws4 = wb["Sheet" + str(n)]                 # accediendo a la Sheetn
    
    ###     PERMUTACIONES     ###
    ws4['E10'] = "Permutaciones"
    perm = permutations(list(range(1,n+1)))    # obtenemos todas las permutaciones de la lista 
    for count, fila in enumerate(list(perm)):  # imprimimos todas las permutaciones
        a = list(fila)                     # la primera fila es una tupla con (1,2,3,4)
        ws4.cell(row=count+11, column=5).value = str(a)
        
        ###     ALGORITMO MEJORADO     ###
        a_original = a[:]    # trabajando con la pila a, y haciendo una copia
        b = []               # la pila b inicialmente está vacía
        b_original = b[:]
        ordenes = []
        while len(a) > 1 and not lista_ordenada(a):   # se repite hasta que solo quede un elemento en la pila a
            # buscamos el mínimo y vamos haciendo ra o rra hasta que el mínimo esté el primero
            while min(a) != a[0]:
                if a.index(min(a)) <= int(len(a)/2):
                    ra(a,b)
                    ordenes.append('ra')
                else:
                    rra(a,b)
                    ordenes.append('rra')
            if not lista_ordenada(a):
                pb(a,b)   # luego hacemos pb
                ordenes.append('pb')
        while len(b) > 0:            # luego se hace pa pa pa pa hasta que no quede nadie en la pila b
            pa(a,b)
            ordenes.append('pa')
        ws4.cell(row=count+11, column=6).value = str(ordenes)
        wb.save('mejor_opcion.xlsx')
        
        ###     FUERZA BRUTA     ###
        catalogo = ["sa","sb","pa","pb","ra","rb","rra","rrb"]  # "ss","rr","rrr"
        mejor_respuesta = [None]*(100+len(a_original)**3)
        las_mejores = []
        contador = 0   # cuenta cuantos grupos de respuestas hay en las_mejores
        ciclos = 0     # cuenta cuantos cliclos da el while True
        while True:
            ciclos += 1
            instrucciones = []
            a = a_original[:]
            b = b_original[:]
            buscando = True
            while not lista_ordenada(a) and buscando:
                instruccion = choice(catalogo)
                instrucciones.append(instruccion)
                if len(instrucciones) > len(mejor_respuesta): buscando = False
                if instruccion=="sa": sa(a,b)
                elif instruccion=="sb": sb(a,b)
                elif instruccion=="ss": ss(a,b)
                elif instruccion=="pa": pa(a,b)
                elif instruccion=="pb": pb(a,b)
                elif instruccion=="ra": ra(a,b)
                elif instruccion=="rb": rb(a,b)
                elif instruccion=="rr": rr(a,b)
                elif instruccion=="rra": rra(a,b)
                elif instruccion=="rrb": rrb(a,b)
                elif instruccion=="rrr": rrr(a,b)
            if len(instrucciones) < len(mejor_respuesta):
                las_mejores = []
            if len(instrucciones) <= len(mejor_respuesta):
                if instrucciones not in las_mejores:
                    mejor_respuesta = instrucciones
                    las_mejores.append(mejor_respuesta)
                    contador += 1
                    ws4.cell(row=count+11, column=6+contador).value = str(mejor_respuesta)
        wb.save('mejor_opcion.xlsx')
